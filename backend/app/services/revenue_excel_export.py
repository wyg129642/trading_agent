"""Export a RevenueModel to an Excel workbook.

Layout mirrors the Modeller-example.xlsx reference: segments grouped in
a table, margin cascade below, with color coding by source_type:

  * yellow  — assumption
  * blue    — guidance
  * green   — historical
  * purple  — expert
  * gray    — derived / inferred

Each cell is annotated with:
  * its formula (as a real Excel formula when possible)
  * a comment containing source notes and citation count
  * the confidence tag in the adjacent column
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from backend.app.models.revenue_model import ModelCell, RevenueModel


_SOURCE_FILL = {
    "historical": PatternFill("solid", fgColor="D1FAE5"),   # green
    "guidance":   PatternFill("solid", fgColor="DBEAFE"),   # blue
    "expert":     PatternFill("solid", fgColor="EDE9FE"),   # purple
    "inferred":   PatternFill("solid", fgColor="F1F5F9"),   # slate
    "assumption": PatternFill("solid", fgColor="FEF3C7"),   # yellow
    "derived":    PatternFill("solid", fgColor="E2E8F0"),   # gray
}


def export_model_to_excel(model: RevenueModel, cells: list[ModelCell]) -> bytes:
    """Build an xlsx bytes payload representing the model."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{model.ticker}"[:30]

    # Header row
    ws["A1"] = f"{model.company_name} ({model.ticker})"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Industry: {model.industry}   Periods: {', '.join(model.fiscal_periods)}"
    ws["A2"].font = Font(color="64748B")

    # Build the table: rows = cells (sorted by path), columns = path, label, value, unit,
    # formula, source_type, confidence, citation count, notes
    headers = ["Path", "Label", "Period", "Value", "Unit", "Formula",
               "Source", "Confidence", "Citations", "Notes"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")
        cell.fill = PatternFill("solid", fgColor="F8FAFC")

    sorted_cells = sorted(cells, key=lambda c: (c.path.split(".")[0], c.path))
    for ri, c in enumerate(sorted_cells, start=5):
        ws.cell(row=ri, column=1, value=c.path).alignment = Alignment(horizontal="left")
        ws.cell(row=ri, column=2, value=c.label)
        ws.cell(row=ri, column=3, value=c.period)
        val_cell = ws.cell(row=ri, column=4)
        if c.value is not None:
            val_cell.value = c.value
            if c.value_type == "percent":
                val_cell.number_format = "0.0%"
            elif c.value_type == "currency":
                val_cell.number_format = "#,##0.00"
            elif c.value_type == "count":
                val_cell.number_format = "#,##0"
        elif c.value_text:
            val_cell.value = c.value_text
        ws.cell(row=ri, column=5, value=c.unit)
        ws.cell(row=ri, column=6, value=c.formula or "")
        ws.cell(row=ri, column=7, value=c.source_type)
        ws.cell(row=ri, column=8, value=c.confidence)
        ws.cell(row=ri, column=9, value=len(c.citations or []))
        ws.cell(row=ri, column=10, value=c.notes[:200] if c.notes else "")

        fill = _SOURCE_FILL.get(c.source_type)
        if fill:
            for col in range(1, 11):
                ws.cell(row=ri, column=col).fill = fill

        # Comment on value cell
        notes_summary = []
        if c.confidence_reason:
            notes_summary.append(f"Why: {c.confidence_reason}")
        if c.citations:
            notes_summary.append(f"{len(c.citations)} source(s): " +
                                 " · ".join(cc.get("title", "") for cc in c.citations[:3]))
        if c.notes:
            notes_summary.append(c.notes[:300])
        if c.alternative_values:
            notes_summary.append(
                f"Alternatives: "
                + " | ".join(f"{a.get('value')} ({a.get('source')})"
                             for a in c.alternative_values[:3])
            )
        if notes_summary:
            val_cell.comment = Comment("\n".join(notes_summary), "Revenue Agent")

    # Column widths
    for col, w in zip("ABCDEFGHIJ", [44, 28, 10, 16, 10, 40, 12, 11, 10, 40]):
        ws.column_dimensions[col].width = w

    # Legend sheet
    legend = wb.create_sheet("Legend")
    legend["A1"] = "Color legend (by source_type)"
    legend["A1"].font = Font(bold=True, size=12)
    for i, (src, fill) in enumerate(_SOURCE_FILL.items(), start=2):
        legend.cell(row=i, column=1, value=src).fill = fill
        legend.cell(row=i, column=2, value={
            "historical": "Actual from 10-K/10-Q",
            "guidance": "Management guidance",
            "expert": "Expert call / interview",
            "inferred": "LLM derivation from context",
            "assumption": "Researcher assumption",
            "derived": "Computed from other cells (formula)",
        }[src])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
